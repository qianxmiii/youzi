#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
解析AD.xlsx文件，生成仓库码表JSON数据
"""

import json
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("请先安装openpyxl: pip install openpyxl")
    sys.exit(1)

def parse_excel_to_json(excel_file, output_file):
    """
    解析Excel文件并生成JSON格式的仓库码表数据
    """
    try:
        # 加载Excel文件
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # 获取表头（假设第一行是表头）
        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value else "")
        
        # 读取数据
        warehouses = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # 跳过空行
                continue
            
            warehouse = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    warehouse[headers[i]] = value if value is not None else ""
            
            if warehouse:  # 只添加非空记录
                warehouses.append(warehouse)
        
        # 生成JSON数据
        data = {
            "warehouses": warehouses,
            "total": len(warehouses)
        }
        
        # 保存为JSON文件
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"成功解析 {len(warehouses)} 条仓库数据")
        print(f"数据已保存到: {output_file}")
        
        return data
        
    except Exception as e:
        print(f"解析Excel文件时出错: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    excel_file = "AD.xlsx"
    output_file = "data/warehouse_data.js"
    
    # 解析Excel并生成JSON
    data = parse_excel_to_json(excel_file, output_file)
    
    # 同时生成JS格式（作为window.warehouseData）
    js_content = f"// 仓库码表数据\nwindow.warehouseData = {json.dumps(data, ensure_ascii=False, indent=2)};"
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(js_content)
    
    print(f"JS格式数据已保存到: {output_file}")

