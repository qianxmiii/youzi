import argparse
import glob
import json
import os
from typing import Any, Dict, List, Optional

import xlrd
from openpyxl import load_workbook
from xlutils.copy import copy as xl_copy


def find_first(pattern: str, base_dir: str) -> str:
    matches = sorted(glob.glob(os.path.join(base_dir, pattern)))
    if not matches:
        raise FileNotFoundError(f"未找到文件: {pattern}")
    return matches[0]


def find_template_xls(base_dir: str) -> str:
    """优先选用含「模板」的 xls，并排除历史生成文件「自动生成」。"""
    all_xls = sorted(glob.glob(os.path.join(base_dir, "*.xls")))
    if not all_xls:
        raise FileNotFoundError(f"未找到模板: {os.path.join(base_dir, '*.xls')}")
    preferred = [p for p in all_xls if "模板" in os.path.basename(p)]
    if preferred:
        return preferred[0]
    candidates = [p for p in all_xls if "自动生成" not in os.path.basename(p)]
    return candidates[0] if candidates else all_xls[0]


def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def safe_div(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if a is None or b is None or b == 0:
        return None
    return a / b


def load_config(config_path: str) -> Dict[str, Any]:
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def build_invoice_header_index(ws, header_row: int) -> Dict[str, int]:
    header_index: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        name = normalize_cell(ws.cell(header_row, col).value)
        if isinstance(name, str) and name != "":
            header_index[name] = col
    return header_index


def get_by_header(ws, row_idx: int, header_index: Dict[str, int], header_name: str) -> Any:
    col = header_index.get(header_name)
    if col is None:
        return ""
    return ws.cell(row_idx, col).value


def is_end_row(ws, row_idx: int, header_index: Dict[str, int], stop_fields: List[str]) -> bool:
    values = [get_by_header(ws, row_idx, header_index, field) for field in stop_fields]
    return all(v is None or str(v).strip() == "" for v in values)


def apply_computed_fields(
    item: Dict[str, Any],
    row_raw: Dict[str, Any],
    computed_config: Dict[str, Dict[str, Any]],
) -> None:
    for target_field, rule in computed_config.items():
        op = rule.get("op")
        if op == "div":
            numerator = to_float(row_raw.get(rule.get("numerator", "")))
            denominator = to_float(row_raw.get(rule.get("denominator", "")))
            value = safe_div(numerator, denominator)
            item[target_field] = "" if value is None else value


def apply_defaults(item: Dict[str, Any], defaults: Dict[str, Any]) -> None:
    for field, default_value in defaults.items():
        value = item.get(field, "")
        if value is None or (isinstance(value, str) and value.strip() == ""):
            item[field] = default_value


def read_invoice_rows(invoice_path: str, config: Dict[str, Any]) -> List[Dict[str, Any]]:
    wb = load_workbook(invoice_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = int(config.get("invoice_header_row", 4))
    data_start_row = int(config.get("invoice_data_start_row", 5))
    stop_fields = config.get("invoice_stop_fields", ["中文品名", "英文品名", "海关编码"])
    mapping = config.get("direct_mapping", {})
    defaults = config.get("defaults", {})
    computed_config = config.get("computed_mapping", {})

    header_index = build_invoice_header_index(ws, header_row)
    missing_headers = sorted(set(mapping.values()) - set(header_index.keys()))
    if missing_headers:
        raise RuntimeError(f"箱单缺少配置中需要的字段: {', '.join(missing_headers)}")

    required_source_headers = set(mapping.values())
    for rule in computed_config.values():
        if rule.get("op") == "div":
            numerator = rule.get("numerator")
            denominator = rule.get("denominator")
            if numerator:
                required_source_headers.add(numerator)
            if denominator:
                required_source_headers.add(denominator)

    missing_headers = sorted(required_source_headers - set(header_index.keys()))
    if missing_headers:
        raise RuntimeError(f"箱单缺少配置中需要的字段: {', '.join(missing_headers)}")

    rows: List[Dict[str, Any]] = []
    row_idx = data_start_row
    while True:
        if is_end_row(ws, row_idx, header_index, stop_fields):
            break

        row_raw: Dict[str, Any] = {}
        for src_header in required_source_headers:
            row_raw[src_header] = normalize_cell(get_by_header(ws, row_idx, header_index, src_header))

        item: Dict[str, Any] = {}
        for target_field, source_header in mapping.items():
            item[target_field] = row_raw.get(source_header, "")

        apply_computed_fields(item, row_raw, computed_config)
        apply_defaults(item, defaults)

        rows.append(item)
        row_idx += 1

    return rows


def fill_template(template_path: str, output_path: str, rows: List[Dict[str, Any]]) -> None:
    rb = xlrd.open_workbook(template_path, formatting_info=True)
    rsh = rb.sheet_by_index(0)
    headers = [str(v).strip() for v in rsh.row_values(1)]

    wb = xl_copy(rb)
    wsh = wb.get_sheet(0)

    start_row = 2  # 第3行开始写数据（0-based）
    for i, item in enumerate(rows):
        target_row = start_row + i
        for col_idx, header in enumerate(headers):
            value = item.get(header, "")
            wsh.write(target_row, col_idx, value)

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="将箱单发票自动转换为产品导入模板。")
    parser.add_argument(
        "--data-dir",
        default=None,
        help="客户箱单与产品模板的目录，默认为本工具下的 samples（可换成你自己的客户资料文件夹）",
    )
    parser.add_argument("--temp-dir", dest="temp_dir", default=None, help="生成文件目录，省略则用仓库根目录下 temp/product_import")
    parser.add_argument("--invoice", default=None, help="箱单发票文件路径（xlsx）")
    parser.add_argument("--template", default=None, help="产品导入模板路径（xls）")
    parser.add_argument("--config", default=None, help="字段配置文件路径（json）")
    parser.add_argument("--output", default=None, help="输出文件路径（xls）")
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.dirname(os.path.dirname(script_dir))
    data_dir = args.data_dir or os.path.join(script_dir, "samples")
    temp_dir = args.temp_dir or os.path.join(repo_root, "temp", "product_import")

    invoice_path = args.invoice or find_first("*.xlsx", data_dir)
    template_path = args.template or find_template_xls(data_dir)
    config_path = args.config or os.path.join(script_dir, "field_mapping.json")

    default_output = os.path.join(temp_dir, "产品导入模版.xls")
    output_path = args.output or default_output

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    config = load_config(config_path)
    rows = read_invoice_rows(invoice_path, config)
    if not rows:
        raise RuntimeError("未读取到有效产品数据，请检查箱单发票内容（默认第5行开始读取）。")

    fill_template(template_path, output_path, rows)
    print(f"已生成: {output_path}")
    print(f"共写入产品行数: {len(rows)}")


if __name__ == "__main__":
    main()
