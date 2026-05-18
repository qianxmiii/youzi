"""码表 Excel 导入 / 模板导出。"""

from __future__ import annotations

import json
import sqlite3
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from youzi_v2.db.code_tables_repository import CodeTablesRepository, validate_table

_CONFIG_PATH = Path(__file__).resolve().parents[1] / "config" / "code_table_excel_columns.json"


def _load_mapping() -> dict[str, Any]:
    with _CONFIG_PATH.open(encoding="utf-8") as f:
        return json.load(f)


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _cell_int(value: Any, default: int = 0) -> int:
    if value is None or str(value).strip() == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _parse_is_active(value: Any, true_tokens: set[str]) -> bool:
    if value is None or str(value).strip() == "":
        return True
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return int(value) != 0
    return str(value).strip() in true_tokens


def _pick_sheet(wb: Any, names: list[str]) -> Any:
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def parse_excel_rows(
    table: str,
    file_path: Path,
) -> tuple[list[tuple[int, dict[str, Any]]], list[dict[str, Any]]]:
    validate_table(table)
    mapping = _load_mapping()
    col_map: dict[str, str] = mapping["columns"]
    true_tokens = set(mapping.get("is_active_true", []))
    sheet_names: list[str] = mapping.get("sheet_preference", [])

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = _pick_sheet(wb, sheet_names)
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], [{"row": 0, "message": "Excel 为空"}]

    header_index: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _cell_str(cell)
        if key and key in col_map:
            header_index[col_map[key]] = idx

    if "code" not in header_index:
        return [], [{"row": 1, "message": "缺少「编码」列（或 code）"}]

    parsed: list[tuple[int, dict[str, Any]]] = []
    errors: list[dict[str, Any]] = []
    excel_row = 1
    for row in rows_iter:
        excel_row += 1
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        def pick(field: str) -> Any:
            i = header_index.get(field)
            if i is None or i >= len(row):
                return None
            return row[i]

        code = _cell_str(pick("code"))
        if not code:
            errors.append({"row": excel_row, "message": "编码为空，已跳过"})
            continue

        payload: dict[str, Any] = {
            "code": code,
            "name_zh": _cell_str(pick("name_zh")) or "",
            "name_en": _cell_str(pick("name_en")) or "",
            "sort_order": _cell_int(pick("sort_order"), 0),
            "is_active": _parse_is_active(pick("is_active"), true_tokens),
        }
        if table == "port_codes":
            pt = _cell_str(pick("port_type")) or "both"
            payload["port_type"] = pt.lower()
        parsed.append((excel_row, payload))

    wb.close()
    return parsed, errors


def import_excel_file(
    repo: CodeTablesRepository,
    table: str,
    file_path: Path,
) -> dict[str, Any]:
    rows, parse_errors = parse_excel_rows(table, file_path)
    created = 0
    updated = 0
    failed = len(parse_errors)
    errors = list(parse_errors)

    for excel_row, payload in rows:
        try:
            _, is_new = repo.upsert_row(table, payload)
            if is_new:
                created += 1
            else:
                updated += 1
        except (ValueError, sqlite3.IntegrityError) as exc:
            failed += 1
            errors.append({"row": excel_row, "message": str(exc)})

    return {
        "table": validate_table(table),
        "created": created,
        "updated": updated,
        "failed": failed,
        "errors": errors[:50],
    }


def build_template_bytes(table: str) -> bytes:
    validate_table(table)
    headers = ["编码", "中文名", "英文名", "排序", "启用"]
    if table == "port_codes":
        headers.insert(3, "港口类型")
    sample = ["SAMPLE01", "示例中文", "Sample EN", 10, "是"]
    if table == "port_codes":
        sample.insert(3, "both")

    wb = Workbook()
    ws = wb.active
    ws.title = "码表"
    ws.append(headers)
    ws.append(sample)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
