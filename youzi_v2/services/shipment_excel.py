"""运单 Excel 解析与导入（对齐运单数据导出模板）。"""

from __future__ import annotations

import json
from pathlib import Path
from typing import TYPE_CHECKING, Any

from io import BytesIO

from openpyxl import Workbook, load_workbook

from youzi_v2.db.shipments_repository import ShipmentsRepository

if TYPE_CHECKING:
    from youzi_v2.db.shipment_groups_repository import ShipmentGroupsRepository
from youzi_v2.internal_tracking import mask_internal_summary

_CONFIG_PATH = Path(__file__).resolve().parents[1] / "config" / "shipment_excel_columns.json"


def _load_mapping() -> dict[str, Any]:
    with _CONFIG_PATH.open(encoding="utf-8") as f:
        return json.load(f)


def _import_column_map(mapping: dict[str, Any]) -> dict[str, str]:
    """导入用：规范表头 + column_aliases + group_import_columns。"""
    col_map: dict[str, str] = dict(mapping["columns"])
    for label, field in mapping.get("column_aliases", {}).items():
        col_map[str(label).strip()] = field
    for label, field in mapping.get("group_import_columns", {}).items():
        col_map[str(label).strip()] = field
    for label, field in mapping.get("group_column_aliases", {}).items():
        col_map[str(label).strip()] = field
    return col_map


def _group_import_fields(mapping: dict[str, Any]) -> set[str]:
    fields: set[str] = set()
    for field in mapping.get("group_import_columns", {}).values():
        fields.add(str(field))
    return fields


def _parse_last_batch(raw: str | None) -> bool:
    text = (raw or "").strip().upper()
    if not text:
        return False
    return text in {"1", "Y", "YES", "TRUE", "是", "最后一批", "末批", "LAST_BATCH", "LAST"}


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _cell_int(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _infer_address_type(
    customer_no: str | None,
    address_code: str | None,
    delivery_address: str | None,
) -> str | None:
    blob = " ".join(
        x for x in (customer_no or "", address_code or "", delivery_address or "") if x
    ).upper()
    if "FBA" in blob or (customer_no or "").upper().startswith("FBA"):
        return "AMZ"
    if "WFS" in blob:
        return "WFS"
    if "3PL" in blob:
        return "3PL"
    return None


def _normalize_country(raw: str | None, aliases: dict[str, str]) -> str | None:
    if not raw:
        return None
    text = raw.strip()
    return aliases.get(text, text)


def _normalize_status(raw: str | None, aliases: dict[str, str]) -> str | None:
    if not raw:
        return None
    text = raw.strip()
    if text in aliases:
        return aliases[text]
    upper = text.upper()
    return upper if upper else None


def _pick_sheet(wb: Any, names: list[str]) -> Any:
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def parse_excel_rows(
    file_path: Path,
) -> tuple[list[tuple[int, dict[str, Any]]], list[dict[str, Any]], list[str]]:
    """
    解析 Excel，返回 (数据行, 行级错误, 已跳过的未映射表头)。
    无法映射的列名仅忽略，不导致整表导入失败。
    """
    mapping = _load_mapping()
    col_map = _import_column_map(mapping)
    export_only: set[str] = set(mapping.get("export_only_fields", []))
    group_fields = _group_import_fields(mapping)
    country_aliases: dict[str, str] = mapping.get("country_aliases", {})
    status_aliases: dict[str, str] = mapping.get("status_aliases", {})
    sheet_names: list[str] = mapping.get("sheet_preference", [])

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = _pick_sheet(wb, sheet_names)
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], [{"row": 1, "message": "空文件"}], []

        header_index: dict[int, str] = {}
        skipped_columns: list[str] = []
        for idx, cell in enumerate(header_row):
            label = _cell_str(cell)
            if not label:
                continue
            if label not in col_map:
                skipped_columns.append(label)
                continue
            field = col_map[label]
            if field in export_only:
                skipped_columns.append(label)
                continue
            header_index[idx] = field

        if "shipment_no" not in header_index.values():
            return [], [{"row": 1, "message": "缺少表头「运单号」"}], skipped_columns

        payloads: list[tuple[int, dict[str, Any]]] = []
        errors: list[dict[str, Any]] = []
        excel_row = 1
        for row in rows_iter:
            excel_row += 1
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            record: dict[str, Any] = {}
            for idx, field in header_index.items():
                val = row[idx] if idx < len(row) else None
                if field == "ctns":
                    parsed = _cell_int(val)
                    if parsed is not None:
                        record[field] = parsed
                else:
                    parsed = _cell_str(val)
                    if parsed is not None:
                        record[field] = parsed

            shipment_no = (record.get("shipment_no") or "").strip()
            if not shipment_no:
                errors.append({"row": excel_row, "message": "运单号为空，已跳过"})
                continue

            if record.get("country_code"):
                record["country_code"] = _normalize_country(
                    record["country_code"], country_aliases
                )
            if record.get("status_code"):
                record["status_code"] = _normalize_status(
                    record["status_code"], status_aliases
                )
            inferred = _infer_address_type(
                record.get("customer_no"),
                record.get("address_code"),
                record.get("delivery_address"),
            )
            if inferred:
                record["address_type"] = inferred

            payloads.append((excel_row, record))
        return payloads, errors, skipped_columns
    finally:
        wb.close()


_API_TO_SNAKE: dict[str, str] = {
    "shipmentNo": "shipment_no",
    "statusCode": "status_code",
    "customerNo": "customer_no",
    "customerShipmentId": "customer_shipment_id",
    "amazonRefId": "amazon_ref_id",
    "customer": "customer",
    "ctns": "ctns",
    "countryCode": "country_code",
    "addressCode": "address_code",
    "deliveryAddress": "delivery_address",
    "zipcode": "zipcode",
    "channelCode": "channel_code",
    "carrierCode": "carrier_code",
    "originWarehouseCode": "origin_warehouse_code",
    "latestTrackingTime": "latest_tracking_time",
    "latestTrackingDesc": "latest_tracking_desc",
    "atd": "atd",
    "deliveredTime": "delivered_time",
}

_STATUS_EXPORT_LABELS: dict[str, str] = {
    "IN_TRANSIT": "转运中",
    "DELIVERED": "已签收",
    "INSPECTION": "查验",
    "UNKNOWN": "未知",
}


def _export_cell_value(field: str, row: dict[str, Any], mapping: dict[str, Any]) -> Any:
    if field in ("latest_tracking_time", "latest_tracking_desc"):
        time_raw = row.get("latest_tracking_time")
        desc_raw = row.get("latest_tracking_desc")
        latest_time, latest_desc = mask_internal_summary(time_raw, desc_raw)
        val = latest_time if field == "latest_tracking_time" else latest_desc
    else:
        val = row.get(field)
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if field == "status_code":
        code = str(val).strip().upper()
        aliases = mapping.get("status_aliases", {})
        for zh, en in aliases.items():
            if en == code:
                return zh
        return _STATUS_EXPORT_LABELS.get(code, val)
    if field == "country_code":
        code = str(val).strip()
        for zh, en in mapping.get("country_aliases", {}).items():
            if en == code:
                return zh
        return val
    if field == "ctns":
        try:
            return int(val)
        except (TypeError, ValueError):
            return val
    if field in ("delivered_time", "atd"):
        text = str(val).strip()
        return text[:10] if len(text) >= 10 else text
    return val


def build_export_excel_bytes(items: list[dict[str, Any]]) -> bytes:
    """按 shipment_excel_columns.json 表头导出，与导入模板一致。"""
    mapping = _load_mapping()
    col_map: dict[str, str] = mapping["columns"]
    headers = list(col_map.keys())
    fields = [col_map[h] for h in headers]

    wb = Workbook()
    ws = wb.active
    sheet_names: list[str] = mapping.get("sheet_preference", [])
    ws.title = (sheet_names[0] if sheet_names else "运单信息")[:31]
    ws.append(headers)

    for item in items:
        row = {_API_TO_SNAKE.get(k, k): v for k, v in item.items()}
        ws.append([_export_cell_value(f, row, mapping) for f in fields])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _split_shipment_and_group_payload(
    record: dict[str, Any],
    group_fields: set[str],
) -> tuple[dict[str, Any], dict[str, Any] | None]:
    shipment = {k: v for k, v in record.items() if k not in group_fields}
    group_no = (record.get("group_no") or "").strip()
    if not group_no:
        return shipment, None
    is_last = _parse_last_batch(str(record.get("is_last_batch") or ""))
    return shipment, {
        "group_no": group_no,
        "group_name": (record.get("group_name") or "").strip(),
        "batch_no": (record.get("batch_no") or "").strip(),
        "role": "LAST_BATCH" if is_last else "NORMAL",
        "customer": (record.get("customer") or "").strip() or None,
        "customer_no": (record.get("customer_no") or "").strip() or None,
    }


def import_excel_file(
    repo: ShipmentsRepository,
    file_path: Path,
    groups_repo: ShipmentGroupsRepository | None = None,
) -> dict[str, Any]:
    mapping = _load_mapping()
    group_fields = _group_import_fields(mapping)
    rows, parse_errors, skipped_columns = parse_excel_rows(file_path)
    created = 0
    updated = 0
    groups_created = 0
    groups_touched = 0
    members_added = 0
    group_errors: list[dict[str, Any]] = []
    errors = list(parse_errors)
    for excel_row, record in rows:
        shipment_payload, group_meta = _split_shipment_and_group_payload(record, group_fields)
        try:
            row, is_new = repo.upsert_by_shipment_no(shipment_payload)
            if is_new:
                created += 1
            else:
                updated += 1
            if groups_repo and group_meta:
                try:
                    group, is_new_group = groups_repo.get_or_create_for_import(
                        group_meta["group_no"],
                        group_name=group_meta.get("group_name") or "",
                        customer=group_meta.get("customer"),
                        customer_no=group_meta.get("customer_no"),
                    )
                    if is_new_group:
                        groups_created += 1
                    else:
                        groups_touched += 1
                    add_res = groups_repo.add_members(
                        group["id"],
                        [row["id"]],
                        role=group_meta.get("role") or "NORMAL",
                        batch_no=group_meta.get("batch_no") or "",
                    )
                    members_added += int(add_res.get("added") or 0)
                except Exception as gexc:  # noqa: BLE001
                    group_errors.append(
                        {
                            "row": excel_row,
                            "message": str(gexc),
                            "shipmentNo": shipment_payload.get("shipment_no"),
                            "groupNo": group_meta.get("group_no"),
                        }
                    )
        except Exception as exc:  # noqa: BLE001 — 汇总行级错误
            errors.append(
                {
                    "row": excel_row,
                    "message": str(exc),
                    "shipmentNo": shipment_payload.get("shipment_no"),
                }
            )
    return {
        "ok": True,
        "totalRows": len(rows),
        "created": created,
        "updated": updated,
        "failed": len(errors),
        "errors": errors[:50],
        "skippedColumns": skipped_columns,
        "groupsCreated": groups_created,
        "groupsTouched": groups_touched,
        "membersAdded": members_added,
        "groupErrors": group_errors[:50],
    }
