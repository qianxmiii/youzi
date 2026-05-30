"""地址簿 Excel 导入测试。"""

from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook

from youzi_v2.db.addresses_warehouse_table import AddressesWarehouseRepository
from youzi_v2.db.connection import Database
from youzi_v2.services.address_excel import (
    build_export_excel_bytes,
    import_excel_file,
    parse_excel_rows,
)


def _write_sample_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "仓库代码",
            "地址类型",
            "收件人公司名",
            "收件人",
            "国家代码",
            "邮编",
            "州省",
            "城市",
            "地址一",
            "地址二",
            "地址三",
            "电话",
            "备注一",
            "备注二",
        ]
    )
    ws.append(
        [
            "TEST1",
            "AMZ",
            "Acme",
            "Alice",
            "US",
            "90001",
            "CA",
            "Los Angeles",
            "100 Main St",
            "",
            "",
            "555-0001",
            "note A",
            "note B",
        ]
    )
    wb.save(path)


def test_parse_excel_rows():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "addresses.xlsx"
        _write_sample_xlsx(path)
        rows, errors = parse_excel_rows(path)
        assert not errors
        assert len(rows) == 1
        _, payload = rows[0]
        assert payload["warehouse_code"] == "TEST1"
        assert payload["address_type"] == "AMZ"
        assert payload["city"] == "Los Angeles"
        assert payload["note1"] == "note A"


def test_import_upsert_by_warehouse_code():
    with tempfile.TemporaryDirectory() as tmp:
        db_path = Path(tmp) / "test.db"
        database = Database(db_path)
        repo = AddressesWarehouseRepository(database)
        path = Path(tmp) / "addresses.xlsx"
        _write_sample_xlsx(path)

        first = import_excel_file(repo, path)
        assert first["created"] == 1
        assert first["updated"] == 0
        assert first["failed"] == 0

        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "仓库代码",
                "地址类型",
                "收件人公司名",
                "收件人",
                "国家代码",
                "邮编",
                "州省",
                "城市",
                "地址一",
                "地址二",
                "地址三",
                "电话",
                "备注一",
                "备注二",
            ]
        )
        ws.append(
            [
                "TEST1",
                "WFS",
                "Acme Updated",
                "Bob",
                "US",
                "90002",
                "CA",
                "LA",
                "200 Main St",
                "",
                "",
                "555-0002",
                "n1",
                "n2",
            ]
        )
        wb.save(path)

        second = import_excel_file(repo, path)
        assert second["created"] == 0
        assert second["updated"] == 1
        assert second["failed"] == 0

        rows = repo.list_rows(search="TEST1")
        assert rows["total"] == 1
        assert len(rows["items"]) == 1
        assert rows["items"][0]["addressType"] == "WFS"
        assert rows["items"][0]["company"] == "Acme Updated"
        assert rows["items"][0]["note2"] == "n2"
        database.conn.close()


def test_parse_excel_rows_accepts_amz_and_maps_fba():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "addresses.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["仓库代码", "地址类型"])
        ws.append(["TEST2", "AMZ"])
        ws.append(["TEST3", "FBA"])
        wb.save(path)

        rows, errors = parse_excel_rows(path)
        assert not errors
        assert len(rows) == 2
        assert rows[0][1]["address_type"] == "AMZ"
        assert rows[1][1]["address_type"] == "AMZ"


def test_build_export_excel_bytes():
    data = build_export_excel_bytes(
        [
            {
                "warehouseCode": "PHX3",
                "addressType": "FBA",
                "company": "Acme",
                "contact": "Alice",
                "countryCode": "US",
                "postalCode": "85043",
                "state": "AZ",
                "city": "Phoenix",
                "addressLine1": "6835 West Buckeye Road",
                "addressLine2": "",
                "addressLine3": "",
                "phone": "555-0001",
                "note1": "备注一",
                "note2": "备注二",
            }
        ]
    )
    assert data[:2] == b"PK"
