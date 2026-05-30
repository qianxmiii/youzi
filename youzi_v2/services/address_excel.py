"""地址簿 Excel 导入 / 模板导出。"""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from youzi_v2.db.addresses_warehouse_table import AddressesWarehouseRepository

_CONFIG_PATH = Path(__file__).resolve().parents[1] / "config" / "address_excel_columns.json"

_TEMPLATE_HEADERS = [
    "仓库代码",
    "地址类型",
    "收件人公司名",
    "收件人",
    "国家代码",
    "邮编",
    "州省",
    "城市",
    "地址一",
    "地址二",
    "地址三",
    "电话",
    "备注一",
    "备注二",
]

_TEMPLATE_SAMPLE = [
    "DFW2",
    "AMZ",
    "Example Corp",
    "John Doe",
    "US",
    "75261",
    "TX",
    "Dallas",
    "1234 Warehouse Blvd",
    "",
    "",
    "+1-555-0100",
    "",
    "",
]


def _load_mapping() -> dict[str, Any]:
    with _CONFIG_PATH.open(encoding="utf-8") as f:
        return json.load(f)


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _pick_sheet(wb: Any, names: list[str]) -> Any:
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def _normalize_address_type(value: Any, mapping: dict[str, Any]) -> str:
    raw = _cell_str(value)
    if not raw:
        return ""
    key = raw.upper()
    aliases: dict[str, str] = mapping.get("address_type_aliases") or {}
    if key in aliases:
        return aliases[key]
    if raw in aliases:
        return aliases[raw]
    return key


def parse_excel_rows(
    file_path: Path,
) -> tuple[list[tuple[int, dict[str, Any]]], list[dict[str, Any]]]:
    mapping = _load_mapping()
    col_map: dict[str, str] = mapping["columns"]
    sheet_names: list[str] = mapping.get("sheet_preference", [])

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = _pick_sheet(wb, sheet_names)
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], [{"row": 0, "message": "Excel 为空"}]

    header_index: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _cell_str(cell)
        if key and key in col_map:
            header_index[col_map[key]] = idx

    if "warehouse_code" not in header_index:
        wb.close()
        return [], [{"row": 1, "message": "缺少「仓库代码」列"}]

    parsed: list[tuple[int, dict[str, Any]]] = []
    errors: list[dict[str, Any]] = []
    excel_row = 1
    for row in rows_iter:
        excel_row += 1
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        def pick(field: str) -> Any:
            i = header_index.get(field)
            if i is None or i >= len(row):
                return None
            return row[i]

        warehouse_code = _cell_str(pick("warehouse_code"))
        if not warehouse_code:
            errors.append({"row": excel_row, "message": "仓库代码为空，已跳过"})
            continue

        payload: dict[str, Any] = {
            "warehouse_code": warehouse_code,
            "address_type": _normalize_address_type(pick("address_type"), mapping),
            "company": _cell_str(pick("company")) or "",
            "contact": _cell_str(pick("contact")) or "",
            "country_code": _cell_str(pick("country_code")) or "",
            "postal_code": _cell_str(pick("postal_code")) or "",
            "state": _cell_str(pick("state")) or "",
            "city": _cell_str(pick("city")) or "",
            "address_line1": _cell_str(pick("address_line1")) or "",
            "address_line2": _cell_str(pick("address_line2")) or "",
            "address_line3": _cell_str(pick("address_line3")) or "",
            "phone": _cell_str(pick("phone")) or "",
            "note1": _cell_str(pick("note1")) or "",
            "note2": _cell_str(pick("note2")) or "",
        }
        parsed.append((excel_row, payload))

    wb.close()
    return parsed, errors


def import_excel_file(
    repo: AddressesWarehouseRepository,
    file_path: Path,
) -> dict[str, Any]:
    rows, parse_errors = parse_excel_rows(file_path)
    created = 0
    updated = 0
    failed = len(parse_errors)
    errors = list(parse_errors)

    for excel_row, payload in rows:
        try:
            _, is_new = repo.upsert_row(payload)
            if is_new:
                created += 1
            else:
                updated += 1
        except ValueError as exc:
            failed += 1
            errors.append({"row": excel_row, "message": str(exc)})

    return {
        "created": created,
        "updated": updated,
        "failed": failed,
        "errors": errors[:50],
    }


def build_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "地址簿"
    ws.append(_TEMPLATE_HEADERS)
    ws.append(_TEMPLATE_SAMPLE)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _item_to_row(item: dict[str, Any]) -> list[Any]:
    return [
        item.get("warehouseCode") or "",
        item.get("addressType") or "",
        item.get("company") or "",
        item.get("contact") or "",
        item.get("countryCode") or "",
        item.get("postalCode") or "",
        item.get("state") or "",
        item.get("city") or "",
        item.get("addressLine1") or "",
        item.get("addressLine2") or "",
        item.get("addressLine3") or "",
        item.get("phone") or "",
        item.get("note1") or "",
        item.get("note2") or "",
    ]


def build_export_excel_bytes(items: list[dict[str, Any]]) -> bytes:
    """导出地址簿，表头与导入模板一致。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "地址簿"
    ws.append(_TEMPLATE_HEADERS)
    for item in items:
        ws.append(_item_to_row(item))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
