"""航次船期 Excel 导入 / 模板导出。"""

from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from youzi_v2.db.vessel_schedules_repository import VesselSchedulesRepository
from youzi_v2.services.vessel_voyage_fields import compose_vessel_voyage

_CONFIG_PATH = Path(__file__).resolve().parents[1] / "config" / "vessel_schedule_excel_columns.json"


def _load_mapping() -> dict[str, Any]:
    with _CONFIG_PATH.open(encoding="utf-8") as f:
        return json.load(f)


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value).strip()
    return text if text else None


def _cell_int(value: Any, default: int) -> int:
    if value is None or str(value).strip() == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _pick_sheet(wb: Any, names: list[str]) -> Any:
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def parse_excel_rows(
    file_path: Path,
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    mapping = _load_mapping()
    col_map: dict[str, str] = mapping["columns"]
    sheet_names: list[str] = mapping.get("sheet_preference", [])

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = _pick_sheet(wb, sheet_names)
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return {}, [{"row": 0, "message": "Excel 为空"}]

    header_index: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _cell_str(cell)
        if key and key in col_map:
            header_index[col_map[key]] = idx

    has_voyage_key = "vessel_voyage" in header_index or (
        "vessel_name" in header_index and "voyage_no" in header_index
    )
    if not has_voyage_key:
        wb.close()
        return {}, [{"row": 1, "message": "缺少「船名航次」列，或同时缺少「船名」「航次」列"}]
    if "port_name" not in header_index:
        wb.close()
        return {}, [{"row": 1, "message": "缺少「港口」列"}]

    grouped: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "port_calls": [],
            "notes": "",
            "vessel_name": None,
            "voyage_no": None,
            "vessel_code": None,
            "shipping_company": None,
        }
    )
    errors: list[dict[str, Any]] = []
    excel_row = 1
    for row in rows_iter:
        excel_row += 1
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        def pick(field: str) -> Any:
            i = header_index.get(field)
            if i is None or i >= len(row):
                return None
            return row[i]

        vessel_voyage = _cell_str(pick("vessel_voyage"))
        vessel_name = _cell_str(pick("vessel_name"))
        voyage_no = _cell_str(pick("voyage_no"))
        voyage_key = vessel_voyage or compose_vessel_voyage(vessel_name, voyage_no)
        port_name = _cell_str(pick("port_name"))
        if not voyage_key:
            errors.append({"row": excel_row, "message": "船名航次为空，已跳过"})
            continue
        if not port_name:
            errors.append({"row": excel_row, "message": "港口为空，已跳过"})
            continue

        notes = _cell_str(pick("notes"))
        if notes:
            grouped[voyage_key]["notes"] = notes
        if vessel_name:
            grouped[voyage_key]["vessel_name"] = vessel_name
        if voyage_no:
            grouped[voyage_key]["voyage_no"] = voyage_no
        vessel_code = _cell_str(pick("vessel_code"))
        if vessel_code:
            grouped[voyage_key]["vessel_code"] = vessel_code
        shipping_company = _cell_str(pick("shipping_company"))
        if shipping_company:
            grouped[voyage_key]["shipping_company"] = shipping_company
        grouped[voyage_key]["port_calls"].append(
            {
                "sequence": _cell_int(pick("sequence"), len(grouped[voyage_key]["port_calls"]) + 1),
                "port_name": port_name,
                "eta": _cell_str(pick("eta")),
                "ata": _cell_str(pick("ata")),
                "etd": _cell_str(pick("etd")),
                "atd": _cell_str(pick("atd")),
            }
        )

    wb.close()
    return dict(grouped), errors


def import_excel_file(
    repo: VesselSchedulesRepository,
    file_path: Path,
) -> dict[str, Any]:
    grouped, parse_errors = parse_excel_rows(file_path)
    created = 0
    updated = 0
    failed = len(parse_errors)
    errors = list(parse_errors)

    for vessel_voyage, payload in grouped.items():
        try:
            _, is_new = repo.upsert_from_import(
                vessel_voyage=vessel_voyage,
                port_calls=payload.get("port_calls") or [],
                notes=str(payload.get("notes") or ""),
                vessel_name=payload.get("vessel_name"),
                voyage_no=payload.get("voyage_no"),
                vessel_code=payload.get("vessel_code"),
                shipping_company=payload.get("shipping_company"),
            )
            if is_new:
                created += 1
            else:
                updated += 1
        except ValueError as exc:
            failed += 1
            errors.append({"row": 0, "message": f"{vessel_voyage}: {exc}"})

    return {
        "created": created,
        "updated": updated,
        "failed": failed,
        "errors": errors[:50],
    }


def build_template_bytes() -> bytes:
    headers = [
        "船名",
        "航次",
        "船舶代码",
        "船公司",
        "船名航次",
        "序号",
        "港口",
        "ETA",
        "ATA",
        "ETD",
        "ATD",
        "备注",
    ]
    samples = [
        [
            "CSCL BOHAI SEA",
            "076E",
            "CSCL-BOHAI-SEA",
            "COSCO",
            "CSCL BOHAI SEA/076E",
            1,
            "Yantian",
            "",
            "",
            "2026-05-10 08:00",
            "2026-05-11 12:00",
            "",
        ],
        [
            "CSCL BOHAI SEA",
            "076E",
            "CSCL-BOHAI-SEA",
            "COSCO",
            "CSCL BOHAI SEA/076E",
            2,
            "Kaohsiung",
            "2026-05-12 10:00",
            "2026-05-12 11:30",
            "2026-05-13 06:00",
            "2026-05-13 08:00",
            "",
        ],
        [
            "CSCL BOHAI SEA",
            "076E",
            "CSCL-BOHAI-SEA",
            "COSCO",
            "CSCL BOHAI SEA/076E",
            3,
            "Long Beach",
            "2026-05-29 16:30",
            "",
            "2026-06-01 10:00",
            "",
            "",
        ],
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "船期"
    ws.append(headers)
    for row in samples:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
